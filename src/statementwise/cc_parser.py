import polars as pl
from abc import ABC, abstractmethod
from openpyxl import load_workbook
import re
from rich import print
from loguru import logger


class CreditCardStatementParser(ABC):
    def __init__(self, file_path, sheet_name=0):
        if not file_path:
            raise ValueError("File path cannot be empty")
        self.file_path = file_path
        self.sheet_name = sheet_name
        self._summary_details = {}
        self.transactions_df = None
    def _get_worksheet(self):
        try:
            workbook = load_workbook(filename=self.file_path)
            if isinstance(self.sheet_name, int):
                sheet = workbook.worksheets[self.sheet_name]
            elif isinstance(self.sheet_name, str):
                if self.sheet_name in workbook.sheetnames:
                    sheet = workbook[self.sheet_name]
                else:
                    raise ValueError(f"Sheet '{self.sheet_name}' not found.")
            else:
                raise ValueError("sheet_name must be an integer index or a string name")
            
            logger.debug("Successfully loaded worksheet")
            return sheet
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found at '{self.file_path}'")
        except Exception as e:
            raise Exception(f"Error loading workbook or sheet: {e}")
        
    @abstractmethod
    def _find_header_row(self, worksheet) -> int:
        pass
    
    @abstractmethod
    def _extract_summary_details(self, worksheet) -> dict:
        pass

    @abstractmethod
    def _map_columns(self, header_row_values:list) -> dict:
        pass
    
    def parse(self):
        
        worksheet = self._get_worksheet()
        
        # 1. get out summary details
        logger.debug(f"[{self.__class__.__name__}] Extracting summary details...")
        self._summary_details = self._extract_summary_details(worksheet)
        logger.debug(f"[{self.__class__.__name__}] Summary details: {self._summary_details}")
        
        2. # detect header row
        logger.debug(f"[{self.__class__.__name__}] Finding header row...")        
        header_row_index = self._find_header_row(worksheet)
        if header_row_index is None:
            raise ValueError(f"Could not find header row for {self.__class__.__name__}")
        logger.debug(f"[{self.__class__.__name__}] Header row found at 0-based index: {header_row_index}")        
        
        logger.debug(f"[{self.__class__.__name__}] Reading transaction data ...")
        try: 
            
            self._transactions_df = pl.read_excel(
                source = self.file_path,
                sheet_name = self.sheet_name,
                read_options={'header_row': header_row_index}
            )
            cols_to_drop = [
                col for col in self._transactions_df.columns
                if self._transactions_df[col].is_null().sum() == self._transactions_df.height
            ]
            if cols_to_drop:
                logger.debug(f"[{self.__class__.__name__}] Dropping entirely null columns: {cols_to_drop}")
                self._transactions_df = self._transactions_df.drop(cols_to_drop)
            
        except Exception as e:
            raise Exception(f"Error reading transaction data with Polars: {e}")
        
        current_headers = self._transactions_df.columns
        column_mapping = self._map_columns(current_headers)
        
        return self._summary_details, self._transactions_df
    
class AmexStatementParser(CreditCardStatementParser):
    AMEX_HEADER_KEYWORDS = [
        'Date', 'Receipt', 'Description', 'Amount', 'Extended Details', 'Apeears On Your Statement As', 'Address', 'City/State', 'Zip Code', 'Country', 'Reference', 'Category'
    ]
    
    AMEX_COLUMN_MAP = {
        'Date': 'transaction_date',
        'Receipt': 'receipt_id',
        'Description': 'description',
        'Amount': 'amount',
        'Extended Details': 'extended_details',
        'Appears On Your Statement As': 'statement_appearance_name',
        'Address': 'merchant_address',
        'City/State': 'merchant_city_state',
        'Zip Code': 'merchant_zip_code',
        'Country': 'merchant_country',
        'Reference': 'amex_reference_number',
        'Category':'category'
    }

    def _find_header_row(self, worksheet) -> int:
        for r_idx in range(worksheet.min_row, min(worksheet.max_row + 1, 51)):
            row_values = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[r_idx]]
            matches = sum(1 for keyword in self.AMEX_HEADER_KEYWORDS if keyword in row_values)
            if matches >= len(self.AMEX_HEADER_KEYWORDS) - 2:
                logger.debug(f"DEBUG: Found header row candidates at 1-based index {r_idx} with {matches} matches. Values: {row_values}")
                return r_idx - 1
        return None
    def _extract_summary_details(self, worksheet) -> dict:
        summary = {}
        
        for r_idx in range(1, min(worksheet.max_row + 1, 10)):
            cell_value_a = str(worksheet.cell(row=r_idx, column=1).value).strip()
            if "Preapred for" in cell_value_a:
                if r_idx + 1 <= worksheet.max_row:
                    prepared_for_value = str(worksheet.cell(row=r_idx+1,column=1).value).strip()
                    summary['prepared_for']=prepared_for_value
            elif "Account Number" in cell_value_a:
                if r_idx + 1 <= worksheet.max_row:
                    account_number_value = str(worksheet.cell(row=r_idx+1,column=1).value).strip()
                    match = re.search(r'(\w{4}-\w{6}-\d{5})', account_number_value)
                    if match:
                        summary['account_number'] = match.group(1)
                    else:
                        summary['account_number'] = account_number_value
        if worksheet.max_row >= 1 and worksheet.max_column >=2:
            a1_value = str(worksheet.cell(row=1, column=1).value).strip()
            b1_value = str(worksheet.cell(row=1, column=2).value).strip()
            
            if "Transaction Details" in a1_value:
                summary['full_statement_header'] = b1_value
                match = re.search(r'^(.*?)(?:\s*\/\s*(.*))?$', b1_value)
                if match:
                    summary['card_type'] = match.group(1).strip()
                    summary['statement_period'] = match.group(2).strip() if match.group(2) else None
                else:
                    summary['card_type'] = b1_value.strip()
                    summary['statement_period'] = None
        return summary
        
                                                                 
    def _map_columns(self, header_row_values: list) -> dict:
        effective_map = {}
        for amex_col_name, standard_name in self.AMEX_COLUMN_MAP.items():
            if amex_col_name in header_row_values:
                effective_map[amex_col_name] = standard_name
        return effective_map